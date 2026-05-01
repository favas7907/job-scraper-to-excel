import argparse
import logging
import os
import re
import sys
import time
from datetime import datetime
from urllib.parse import urljoin

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  [%(levelname)s]  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

API_BASE = "https://remoteok.com/api"
HTML_BASE = "https://remoteok.com"
COMPANY_NAME = "RemoteOK"
REQUEST_DELAY = 1.5
REQUEST_TIMEOUT = 20
MAX_RETRIES = 3

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "application/json, text/html",
}

SKILL_KEYWORDS = [
    "Python", "JavaScript", "TypeScript", "React", "Node.js", "Vue", "Angular",
    "Next.js", "Django", "FastAPI", "Flask", "SQL", "PostgreSQL", "MySQL",
    "MongoDB", "Redis", "Docker", "Kubernetes", "AWS", "GCP", "Azure",
    "REST", "GraphQL", "CI/CD", "Git", "Linux", "Machine Learning",
    "TensorFlow", "PyTorch", "Pandas", "NumPy", "Spark", "Kafka", "Airflow",
    "Go", "Java", "Scala", "Rust", "Ruby", "PHP", "Swift", "Kotlin",
]

EXCEL_COLUMNS = [
    "JobTitle",
    "Location",
    "ExperienceRequired",
    "SkillsRequired",
    "Salary",
    "JobURL",
    "JobDescriptionSummary",
]


def safe_get(data: dict, *keys, fallback: str = "") -> str:
    val = data
    for key in keys:
        if not isinstance(val, dict):
            return fallback
        val = val.get(key)
        if val is None:
            return fallback
    return str(val).strip() if val else fallback


def create_session() -> requests.Session:
    session = requests.Session()
    session.headers.update(HEADERS)
    return session


def fetch_url(url: str, session: requests.Session, as_json: bool = False):
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            logger.info("Fetching (attempt %d): %s", attempt, url)
            r = session.get(url, timeout=REQUEST_TIMEOUT)
            r.raise_for_status()
            if as_json:
                return r.json()
            return BeautifulSoup(r.text, "html.parser")
        except requests.exceptions.HTTPError as e:
            code = e.response.status_code
            logger.warning("HTTP %d for %s", code, url)
            if code in {403, 404, 410}:
                return None
            time.sleep(REQUEST_DELAY * attempt)
        except requests.exceptions.RequestException as e:
            logger.warning("Request error (attempt %d): %s", attempt, e)
            time.sleep(REQUEST_DELAY * attempt)
        except ValueError:
            logger.error("JSON decode error for %s", url)
            return None
    logger.error("All %d attempts failed: %s", MAX_RETRIES, url)
    return None


def strip_html(raw: str) -> str:
    if not raw:
        return ""
    return BeautifulSoup(raw, "html.parser").get_text(separator=" ", strip=True)


def extract_skills(text: str, api_tags: list = None) -> str:
    skills_from_tags = []
    if api_tags and isinstance(api_tags, list):
        skills_from_tags = [t.strip().title() for t in api_tags if t.strip()]

    skills_from_desc = [
        kw for kw in SKILL_KEYWORDS
        if text and re.search(rf"\b{re.escape(kw)}\b", text, re.IGNORECASE)
    ]

    seen, merged = set(), []
    for s in skills_from_tags + skills_from_desc:
        if s.lower() not in seen:
            seen.add(s.lower())
            merged.append(s)

    return ", ".join(merged[:12])


def extract_experience(text: str) -> str:
    if not text:
        return ""
    m = re.search(
        r"(\d+\+?\s*(?:to\s*\d+\s*)?years?\s*(?:of\s*)?(?:experience|exp)?)",
        text,
        re.IGNORECASE,
    )
    return m.group(1).strip() if m else ""


def extract_salary(raw: dict, text: str = "") -> str:
    s_min = raw.get("salary_min")
    s_max = raw.get("salary_max")
    if s_min and s_max:
        return f"${int(s_min):,} – ${int(s_max):,}"
    if s_min:
        return f"${int(s_min):,}+"

    if text:
        m = re.search(
            r"(\$[\d,]+(?:\s*[-–]\s*\$[\d,]+)?(?:\s*(?:k|K|USD|/yr|annually))?)",
            text,
        )
        if m:
            return m.group(1).strip()
    return ""


def fetch_jobs_from_api(tag: str, session: requests.Session) -> list:
    url = f"{API_BASE}?tag={tag}"
    data = fetch_url(url, session, as_json=True)
    if not data or not isinstance(data, list):
        return []
    jobs = [item for item in data if item.get("id") and item.get("position")]
    logger.info("API → %d raw jobs for tag '%s'", len(jobs), tag)
    return jobs


def handle_pagination(soup: BeautifulSoup, current_url: str) -> str | None:
    tag = soup.find("a", rel="next")
    if tag and tag.get("href"):
        return urljoin(HTML_BASE, tag["href"])

    for a in soup.find_all("a", href=True):
        if re.search(r"next|›|»", a.get_text(strip=True), re.IGNORECASE):
            return urljoin(HTML_BASE, a["href"])

    pagination = soup.select_one(".pagination, nav[aria-label='pagination']")
    if pagination:
        active = pagination.find("li", class_=re.compile(r"active|current", re.I))
        if active:
            sibling = active.find_next_sibling("li")
            if sibling:
                link = sibling.find("a", href=True)
                if link:
                    return urljoin(HTML_BASE, link["href"])

    return None


def fetch_jobs_from_html(tag: str, session: requests.Session, max_pages: int = 5) -> list:
    raw_jobs = []
    page_url = f"{HTML_BASE}/remote-{tag}-jobs"
    page_num = 1

    while page_url and page_num <= max_pages:
        logger.info("HTML fallback — page %d: %s", page_num, page_url)
        soup = fetch_url(page_url, session, as_json=False)
        if soup is None:
            break

        cards = soup.select("section.jobs ul.jobs li:not(.view-all)")
        logger.info("  Found %d job cards on page %d", len(cards), page_num)

        for card in cards:
            link = card.find("a", href=True)
            if not link:
                continue
            title_tag = card.select_one("span.title")
            region_tag = card.select_one("span.region, .region, .location")
            raw_jobs.append({
                "position": title_tag.get_text(strip=True) if title_tag else "",
                "location": region_tag.get_text(strip=True) if region_tag else "",
                "url": urljoin(HTML_BASE, link["href"]),
                "tags": [],
                "description": "",
                "salary_min": None,
                "salary_max": None,
            })

        next_url = handle_pagination(soup, page_url)
        page_url = next_url
        page_num += 1
        if next_url:
            time.sleep(REQUEST_DELAY)

    logger.info("HTML fallback collected %d raw jobs", len(raw_jobs))
    return raw_jobs


def extract_job_details(raw: dict) -> dict:
    try:
        title = safe_get(raw, "position")
        location = safe_get(raw, "location") or "Remote – Worldwide"
        job_url = safe_get(raw, "url")
        description = strip_html(safe_get(raw, "description"))
        summary = description[:300].strip()
        salary = extract_salary(raw, description)
        skills = extract_skills(description, raw.get("tags", []))
        experience = extract_experience(description)

        if job_url and not job_url.startswith("http"):
            job_url = urljoin(HTML_BASE, job_url)

        return {
            "JobTitle": title,
            "Location": location,
            "ExperienceRequired": experience,
            "SkillsRequired": skills,
            "Salary": salary,
            "JobURL": job_url,
            "JobDescriptionSummary": summary,
        }
    except Exception as exc:
        logger.warning("extract_job_details error (skipping row): %s", exc)
        return {col: "" for col in EXCEL_COLUMNS}


def parse_jobs(raw_jobs: list, max_results: int) -> list:
    jobs, seen = [], set()

    for raw in raw_jobs:
        if len(jobs) >= max_results:
            break

        job = extract_job_details(raw)
        if not job["JobTitle"]:
            continue
        if job["JobURL"] in seen:
            logger.debug("Duplicate skipped: %s", job["JobURL"])
            continue

        seen.add(job["JobURL"])
        jobs.append(job)
        logger.info(
            "  ✓  %-45s | %-22s | %s",
            job["JobTitle"][:45],
            job["Location"][:22],
            job["Salary"] or "—",
        )
        time.sleep(0.05)

    return jobs


def style_excel(filepath: str) -> None:
    wb = load_workbook(filepath)
    ws = wb.active
    ws.title = "Jobs"

    header_fill = PatternFill("solid", start_color="1F3864", end_color="1F3864")
    row_even = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
    row_odd = PatternFill("solid", start_color="EEF2FF", end_color="EEF2FF")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    cell_font = Font(name="Arial", size=10)
    cell_border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = cell_border
    ws.row_dimensions[1].height = 32

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = row_even if row_idx % 2 == 0 else row_odd
        for cell in row:
            cell.font = cell_font
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = cell_border

    col_widths = {
        "JobTitle": 38,
        "Location": 24,
        "ExperienceRequired": 20,
        "SkillsRequired": 48,
        "Salary": 24,
        "JobURL": 52,
        "JobDescriptionSummary": 62,
    }
    for idx, col_name in enumerate(EXCEL_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = col_widths.get(col_name, 20)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(filepath)
    logger.info("Formatting applied → %s", filepath)


def save_to_excel(jobs: list) -> tuple:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{COMPANY_NAME}_Jobs_{timestamp}.xlsx"

    df = pd.DataFrame(jobs)
    for col in EXCEL_COLUMNS:
        if col not in df.columns:
            df[col] = ""

    df = df[EXCEL_COLUMNS].fillna("")
    df.to_excel(filename, index=False, engine="openpyxl")
    logger.info("Wrote %d rows → %s", len(df), filename)

    style_excel(filename)
    return filename, df


def run_self_test() -> None:
    print("\n" + "═" * 70)
    print("  SELF-TEST — Verifying all requirements offline")
    print("═" * 70)

    test_raw_jobs = [
        {
            "id": "1", "position": "Senior Python Developer",
            "location": "Remote – US Only",
            "url": "https://remoteok.com/jobs/1",
            "tags": ["python", "django", "postgresql"],
            "description": "<p>We need 5+ years of experience with Python, Django, and AWS. "
                           "Salary $120,000 – $150,000. Docker and CI/CD required.</p>",
            "salary_min": 120000, "salary_max": 150000,
        },
        {
            "id": "2", "position": "React Frontend Engineer",
            "location": "Remote – Europe",
            "url": "https://remoteok.com/jobs/2",
            "tags": ["react", "typescript"],
            "description": "<p>3+ years experience. TypeScript and Node.js required.</p>",
            "salary_min": None, "salary_max": None,
        },
        {
            "id": "3", "position": "Data Engineer",
            "location": "",
            "url": "https://remoteok.com/jobs/3",
            "tags": ["python", "spark", "kafka"],
            "description": "<p>Spark, Kafka, Airflow, SQL. 4+ years experience in data pipelines.</p>",
            "salary_min": 95000, "salary_max": None,
        },
        {
            "id": "4", "position": "DevOps Engineer",
            "location": "Remote – APAC",
            "url": "https://remoteok.com/jobs/4",
            "tags": ["devops", "kubernetes", "aws"],
            "description": "",
            "salary_min": None, "salary_max": None,
        },
        {
            "id": "5", "position": "Machine Learning Engineer",
            "location": "Remote – Worldwide",
            "url": "https://remoteok.com/jobs/5",
            "tags": [],
            "description": "<p>Deep knowledge of TensorFlow, PyTorch, Python, and Docker. 6+ years exp.</p>",
            "salary_min": 140000, "salary_max": 170000,
        },
        {
            "id": "1-dup", "position": "Senior Python Developer (duplicate)",
            "location": "Remote – US Only",
            "url": "https://remoteok.com/jobs/1",
            "tags": [], "description": "", "salary_min": None, "salary_max": None,
        },
        {
            "id": "6", "position": "",
            "location": "Remote – Worldwide",
            "url": "https://remoteok.com/jobs/6",
            "tags": [], "description": "Some text.", "salary_min": None, "salary_max": None,
        },
    ]

    PASS = "  ✅ PASS"
    FAIL = "  ❌ FAIL"
    results = []

    try:
        jobs = parse_jobs(test_raw_jobs, max_results=100)
        results.append((PASS, "Req 1 — Script runs without errors"))
    except Exception as e:
        results.append((FAIL, f"Req 1 — Crashed: {e}"))
        jobs = []

    try:
        filename, df = save_to_excel(jobs)
        missing_cols = [c for c in EXCEL_COLUMNS if c not in df.columns]
        if missing_cols:
            results.append((FAIL, f"Req 2 — Missing columns: {missing_cols}"))
        else:
            results.append((PASS, f"Req 2 — Excel has all {len(EXCEL_COLUMNS)} required columns"))
    except Exception as e:
        results.append((FAIL, f"Req 2 — Excel save error: {e}"))
        filename, df = None, pd.DataFrame()

    if not df.empty:
        row = df[df["JobTitle"] == "Senior Python Developer"].iloc[0]
        if "Python" in row["SkillsRequired"] and row["JobURL"]:
            results.append((PASS, "Req 3 — Data accurately reflects source (title, skills, URL)"))
        else:
            results.append((FAIL, "Req 3 — Data mismatch"))

    if not df.empty:
        checks = []
        no_salary = df[df["JobTitle"] == "React Frontend Engineer"]
        if not no_salary.empty and no_salary.iloc[0]["Salary"] == "":
            checks.append("salary blank ✓")
        no_loc = df[df["JobTitle"] == "Data Engineer"]
        if not no_loc.empty and no_loc.iloc[0]["Location"] == "Remote – Worldwide":
            checks.append("location default ✓")
        no_desc = df[df["JobTitle"] == "DevOps Engineer"]
        if not no_desc.empty and no_desc.iloc[0]["JobDescriptionSummary"] == "":
            checks.append("description blank ✓")
        if len(df) == 5:
            checks.append("duplicate excluded ✓")
        if df[df["JobTitle"] == ""].empty:
            checks.append("empty title excluded ✓")

        if len(checks) == 5:
            results.append((PASS, f"Req 4 — Missing data graceful: {', '.join(checks)}"))
        else:
            results.append((FAIL, f"Req 4 — Some checks failed. Passed: {checks}"))

    try:
        dummy_soup = BeautifulSoup('<a rel="next" href="/remote-jobs?page=2">Next</a>', "html.parser")
        next_url = handle_pagination(dummy_soup, "https://remoteok.com/remote-jobs")
        if next_url and "page=2" in next_url:
            results.append((PASS, "Req 5 — Pagination detection works correctly"))
        else:
            results.append((FAIL, "Req 5 — Pagination returned wrong URL"))
    except Exception as e:
        results.append((FAIL, f"Req 5 — Pagination error: {e}"))

    print()
    for status, message in results:
        print(f"{status}  {message}")

    all_passed = all(r[0] == PASS for r in results)
    print()
    if all_passed:
        print("  🎉  ALL REQUIREMENTS VERIFIED SUCCESSFULLY")
    else:
        print("  ⚠️   Some checks failed — see above")

    if filename and os.path.exists(filename):
        print(f"\n  📄  Test Excel file: {filename}")
        print("\n  ─── DataFrame preview ───────────────────────────────────────")
        pd.set_option("display.max_columns", None)
        pd.set_option("display.max_colwidth", 35)
        pd.set_option("display.width", 140)
        print(df[["JobTitle", "Location", "Salary", "ExperienceRequired"]].to_string(index=False))

    print("\n" + "═" * 70 + "\n")


def scrape(query: str, max_results: int) -> list:
    session = create_session()
    raw_jobs = fetch_jobs_from_api(tag=query, session=session)

    if not raw_jobs:
        logger.warning("API returned no data. Falling back to HTML scraping with pagination...")
        raw_jobs = fetch_jobs_from_html(tag=query, session=session, max_pages=5)

    if not raw_jobs:
        logger.error("No jobs found via API or HTML. Check query or connection.")
        return []

    return parse_jobs(raw_jobs, max_results=max_results)


def parse_args():
    parser = argparse.ArgumentParser(
        description="RemoteOK Job Scraper — exports live job data to Excel.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python scraper.py
  python scraper.py --query "react" --max 30
  python scraper.py --query "devops" --max 20
  python scraper.py --selftest
        """,
    )
    parser.add_argument("--query", "-q", default="python", help="Job tag to search (default: python)")
    parser.add_argument("--max", "-m", type=int, default=50, help="Max jobs to collect (default: 50)")
    parser.add_argument("--selftest", action="store_true", help="Run offline requirement verification and exit")
    return parser.parse_args()


def main():
    args = parse_args()

    if args.selftest:
        run_self_test()
        sys.exit(0)

    logger.info("=" * 60)
    logger.info("  RemoteOK Job Scraper")
    logger.info("  Query : %s", args.query)
    logger.info("  Max   : %d jobs", args.max)
    logger.info("=" * 60)

    jobs = scrape(query=args.query, max_results=args.max)

    if not jobs:
        logger.warning("No jobs collected. Exiting.")
        sys.exit(1)

    output_file, df = save_to_excel(jobs)

    print("\n" + "═" * 110)
    print("  SAMPLE OUTPUT — First 5 rows")
    print("═" * 110)
    pd.set_option("display.max_columns", None)
    pd.set_option("display.max_colwidth", 38)
    pd.set_option("display.width", 130)
    print(df[["JobTitle", "Location", "Salary", "ExperienceRequired", "SkillsRequired"]].head())
    print("═" * 110)
    print(f"\n✅  {len(jobs)} jobs saved  →  {output_file}\n")


if __name__ == "__main__":
    main()
